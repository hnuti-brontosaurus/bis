import logging

import requests
from openpyxl import load_workbook


def run():
    wb = load_workbook("/home/lamanchy/laman/Downloads/Uživatelé_f1r26axq.xlsx")
    ws = wb.active

    i = 2
    while user := ws[f"A{i}"].value:
        response = requests.get(
            f"https://bis.brontosaurus.cz/api/frontend/users/{user}/participated_in_events",
            {"page_size": 1000},
            headers={"Authorization": "Token 9bd09b53512d2dfd41d81ee3d16fc12434c505ea"},
        )
        events = response.json()["results"]

        events = [_ for _ in events if "2024" in _["start"]]

        events = "\n".join(f"{_['name']} - {_['start']}" for _ in events)
        ws[f"AG{i}"] = events

        i += 1
        logging.info("Processing row %d", i)

    wb.save("file_modified.xlsx")


if __name__ == "__main__":
    run()
