from datetime import date, timedelta
from tempfile import NamedTemporaryFile

import openpyxl
from administration_units.models import AdministrationUnit
from categories.models import RoleCategory
from dateutil.utils import today
from django.contrib.gis.db.models import *
from django.core.files import File
from donations.models import Donation
from tinymce.models import HTMLField
from translation.translate import translate_model

from bis.helpers import AgeStats, filter_queryset_with_multiple_or_queries
from bis.models import User


@translate_model
class DuplicateUser(Model):
    user = ForeignKey(User, on_delete=CASCADE, related_name="duplicates")
    other = ForeignKey(User, on_delete=CASCADE, related_name="other_duplicates")

    def save(
        self, force_insert=False, force_update=False, using=None, update_fields=None
    ):
        if self.user == self.other:
            if not self._state.adding:
                self.delete()

            return

        super().save(force_insert, force_update, using, update_fields)

    @classmethod
    def filter_queryset(cls, queryset, perm):
        visible_users = User.filter_queryset(User.objects.all(), perm)
        return queryset.filter(user__in=visible_users, other__in=visible_users)

    def can_be_merged_by(self, user):
        if user.is_superuser:
            return True
        if user.is_office_worker:
            return not (self.user.is_superuser or self.other.is_superuser)
        return False

    def __str__(self):
        return "Duplicita"

    class Meta:
        ordering = ("id",)
        unique_together = "user", "other"


@translate_model
class DashboardItem(Model):
    date = DateField()
    visible_date = HTMLField(max_length=63, blank=True)
    name = HTMLField(max_length=63)
    description = HTMLField(blank=True)
    repeats_every_year = BooleanField(default=False)

    for_roles = ManyToManyField(RoleCategory, related_name="dashboard_items")

    def __str__(self):
        return self.name

    class Meta:
        ordering = ("-date",)
        Index(fields=["date"])

    @classmethod
    def get_items_for_user(cls, user):
        dashboard_items = list(
            DashboardItem.objects.filter(
                for_roles__in=user.roles.all(), date__gte=today().date()
            ).distinct()
        )

        for application in user.applications.filter(
            event_registration__event__start__gte=today(),
            event_registration__event__is_canceled=False,
        ):
            event = application.event_registration.event
            dashboard_items.append(
                DashboardItem(date=event.start, name=f"Začíná ti akce {event.name}")
            )

        future_events = []
        for event in (
            user.events_where_was_organizer.exclude(is_archived=True)
            .exclude(is_canceled=True)
            .filter(start__gte=today())
        ):
            dashboard_items.append(
                DashboardItem(date=event.start, name=f"Organizuješ akci {event.name}")
            )
            future_events.append(event)

        for event in (
            user.events_where_was_organizer.exclude(is_archived=True)
            .exclude(is_canceled=True)
            .filter(is_closed=False)
        ):
            if event in future_events:
                continue

            dashboard_items.append(
                DashboardItem(
                    date=event.start + timedelta(days=20),
                    name=f"Deadline pro uzavření akce {event.name}",
                )
            )

        dashboard_items.sort(key=lambda obj: obj.date)

        return dashboard_items


@translate_model
class DonationPointsAggregation(Model):
    name = CharField(max_length=127)
    slug = SlugField(unique=True)
    description = TextField()

    class Meta:
        ordering = ("id",)

    def __str__(self):
        return self.name

    @staticmethod
    def do_count_events(events):
        return sum(events.values_list("number_of_sub_events", flat=True))

    @classmethod
    def do_get_count(cls, slug, since, till, administration_unit):
        donations = Donation.objects.filter(donated_at__gte=since, donated_at__lte=till)
        if slug.startswith("supporting_donations"):
            if slug == "supporting_donations":
                donations = donations.filter(
                    donor__basic_section_support=administration_unit
                )
            if slug == "supporting_donations_rc":
                donations = donations.filter(
                    donor__regional_center_support=administration_unit
                )
            return sum(donations.values_list("amount", flat=True))

        events = administration_unit.events.filter(start__gte=since, start__lte=till)
        if slug == "clubs":
            return cls.do_count_events(
                events.filter(
                    group__slug="other",
                    category__slug__in=[
                        "public__club__meeting",
                        "public__club__lecture",
                        "public__educational__lecture",
                    ],
                )
            )
        if slug == "other_without_clubs":
            return cls.do_count_events(
                events.filter(
                    group__slug="other",
                    category__slug__in=[
                        "public__volunteering",
                        "public__only_experiential",
                        "public__educational__course",
                        "public__other__for_public",
                        "public__other__eco_tent",
                    ],
                )
            )
        if slug == "weekend_events":
            return cls.do_count_events(
                events.filter(
                    group__slug="weekend_event",
                    category__slug__in=[
                        "public__volunteering",
                        "public__only_experiential",
                        "public__educational__course",
                        "public__educational__educational_with_stay",
                        "public__other__for_public",
                        "public__other__eco_tent",
                    ],
                )
            )
        if slug == "camps":
            return cls.do_count_events(
                events.filter(
                    group__slug="camp",
                    category__slug__in=[
                        "public__volunteering",
                        "public__only_experiential",
                        "public__educational__course",
                        "public__educational__educational_with_stay",
                        "public__other__eco_tent",
                    ],
                )
            )

        if slug == "50_worked_hours":
            hours_worked = events.values_list("record__total_hours_worked", flat=True)
            hours_worked = [hours for hours in hours_worked if hours]
            return int(sum(hours_worked) / 50)

        members = administration_unit.memberships.filter(year=till.year)
        users = User.objects.filter(id__in=members.values_list("user_id"))
        age_stats = AgeStats("", users, till)
        if slug == "members_0_15":
            return age_stats.age_count(0, 15)
        if slug == "members_16_18":
            return age_stats.age_count(16, 18)
        if slug == "members_19_26":
            return age_stats.age_count(19, 26)
        if slug == "members_27_and_more":
            return age_stats.age_count(27, age_stats.oldest)

        raise "unknown slug"

    def get_count(self, since, till, administration_unit):
        return self.do_get_count(self.slug, since, till, administration_unit)


@translate_model
class DonationPoints(Model):
    name = CharField(max_length=255)
    since = DateField()
    till = DateField()
    file = FileField(upload_to="donation_points", blank=True)

    def __str__(self):
        return self.name

    def get_donation_points(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        ws.cell(1, 1).value = "Body dotací"
        ws.cell(2, 1).value = "od"
        ws.cell(2, 2).value = str(self.since)
        ws.cell(2, 3).value = "do"
        ws.cell(2, 4).value = str(self.till)

        for r, row in enumerate(self.get_rows()):
            for c, cell in enumerate(row):
                ws.cell(4 + r, 1 + c).value = cell

        tmp_file = NamedTemporaryFile(
            mode="w",
            suffix=".xlsx",
            newline="",
            encoding="utf8",
            prefix="donation_points_",
        )
        wb.save(tmp_file.name)
        return open(tmp_file.name, "rb")

    def save(
        self, force_insert=False, force_update=False, using=None, update_fields=None
    ):
        if not self._state.adding:
            file = File(self.get_donation_points())
            self.file.save(f"{self.name}.xlsx", file, save=False)
        super().save(force_insert, force_update, using, update_fields)

    def get_header(self):
        header = ["Organizační jednotky", "Kategorie", ""]
        for section in self.sections.all():
            header += section.get_header()
        header += ["Celkem"]
        return header

    def get_subheader(self):
        header = ["", "", ""]
        for section in self.sections.all():
            header += section.get_subheader()
        return header

    def get_row(self, administration_unit):
        row = [administration_unit.abbreviation, administration_unit.category.name, ""]
        total = 0
        for section in self.sections.all():
            values, row_total = section.get_row(administration_unit)
            row += values
            total += row_total
        row += [total]
        return row

    def get_rows(self):
        yield self.get_header()
        yield self.get_subheader()

        administration_units = AdministrationUnit.objects.filter(
            existed_till__isnull=True,
        ).order_by("category", "abbreviation", "id")
        for administration_unit in administration_units:
            yield self.get_row(administration_unit)


@translate_model
class DonationPointsSection(Model):
    donation_points = ForeignKey(
        DonationPoints, on_delete=CASCADE, related_name="sections"
    )

    def __str__(self):
        return f"Sekce"

    def get_header(self):
        return [
            f"{column.aggregation.name} - {column.points_per_each} bodů"
            for column in self.columns.all()
        ] + ["Celkem za sekci", ""]

    def get_subheader(self):
        return [column.aggregation.description for column in self.columns.all()] + [
            "",
            "",
        ]

    def get_row(self, administration_unit):
        row = []
        total = 0
        for column in self.columns.all():
            count = column.aggregation.get_count(
                self.donation_points.since,
                self.donation_points.till,
                administration_unit,
            )
            count *= column.points_per_each
            row.append(count)
            total += count

        row += [total, ""]
        return row, total


@translate_model
class DonationPointsColumn(Model):
    aggregation = ForeignKey(DonationPointsAggregation, on_delete=PROTECT)
    points_per_each = FloatField()
    section = ForeignKey(
        DonationPointsSection, on_delete=CASCADE, related_name="columns"
    )

    def __str__(self):
        return f"Sloupec {self.aggregation}"


class SavedFile(Model):
    name = CharField(max_length=63)
    file = FileField(upload_to="saved_files")
    created_at = DateField(auto_now_add=True)

    @classmethod
    def remove_old(cls):
        SavedFile.objects.filter(created_at__lte=today() - timedelta(days=14)).delete()
