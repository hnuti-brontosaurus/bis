import logging
import os
from base64 import b64encode
from collections import Counter
from concurrent.futures import ThreadPoolExecutor
from copy import copy
from datetime import date
from itertools import zip_longest
from os.path import join
from pathlib import Path
from shutil import copy2, make_archive
from tempfile import TemporaryDirectory
from threading import Lock
from time import sleep
from typing import OrderedDict
from zipfile import ZIP_DEFLATED, ZipFile

import openpyxl
import pdfkit
import xlsxwriter
from bis.helpers import print_progress
from bis.models import User
from common.thumbnails import get_thumbnail_path
from django.conf import settings
from django.contrib import admin, messages
from django.core.files.temp import NamedTemporaryFile
from django.core.paginator import Paginator
from django.db.models import Count
from django.http import FileResponse
from django.utils.formats import date_format
from event.models import Event
from PIL import Image, ImageDraw, ImageFont
from project.settings import BASE_DIR
from questionnaire.models import EventApplication
from rest_framework.serializers import ModelSerializer
from translation.translate import _
from xlsx2html import xlsx2html
from xlsx_export.helpers import text_into_lines
from xlsx_export.serializers import (
    AdministrationUnitExportSerializer,
    DonationExportSerializer,
    DonorExportSerializer,
    EventApplicationExportSerializer,
    EventExportSerializer,
    EventFeedbackExportSerializer,
    MembershipExportSerializer,
    UserExportSerializer,
)

from bis import emails

lock = Lock()


class XLSXWriter:
    def __init__(self, file_name):
        self.tmp_file = NamedTemporaryFile(
            mode="w",
            suffix=".xlsx",
            newline="",
            encoding="utf8",
            prefix=file_name + "_",
        )
        self.writer = xlsxwriter.Workbook(self.tmp_file.name, {"constant_memory": True})

        self.format = lambda: None
        self.format.green = self.writer.add_format({"bg_color": "#c9ffc9"})
        self.format.red = self.writer.add_format({"bg_color": "#ff9999"})
        self.format.shrink = self.writer.add_format()
        self.format.shrink.set_shrink()
        self.format.text_wrap = self.writer.add_format()
        self.format.text_wrap.set_text_wrap()
        self.format.bold = self.writer.add_format()
        self.format.bold.set_shrink()
        self.format.bold.set_bold()

    def get_file(self):
        self.writer.close()
        self.tmp_file.flush()

        return self.tmp_file

    def add_worksheet(self, name):
        self.worksheet = self.writer.add_worksheet(name)
        self.row = 0
        self.header_keys = []
        self.widths = {}

    def from_queryset(self, queryset, serializer_class, name=None):
        if name is None:
            name = queryset.model._meta.verbose_name_plural

        self.add_worksheet(name)

        queryset = serializer_class.get_related(queryset)

        for page in Paginator(queryset, 100):
            print_progress("exporting xlsx", page.number, page.paginator.num_pages)
            serializer = serializer_class(page.object_list, many=True)
            for item in serializer.data:
                if not self.row:
                    self.write_header(self.get_fields(serializer, queryset))
                self.write_row(item)

        for i, key in enumerate(self.header_keys):
            self.worksheet.set_column(i, i, width=self.widths[key])

    def get_fields(self, serializer, queryset):
        fields = serializer.child.get_fields()
        if fn := getattr(serializer.child, "get_extra_fields", None):
            fields.update(fn(queryset))
        return fields

    def write_values(self, values):
        values = {key: value for key, value in values}
        height = 0
        for i, key in enumerate(self.header_keys):
            value = values.get(key)
            if isinstance(value, list):
                value = "\n".join(str(v) for v in value)
            if value is False:
                value = "ne"
            if value is True:
                value = "ano"
            if value is None:
                value = "-"
            if not isinstance(value, (str, int, float)):
                value = str(value)
            self.widths.setdefault(key, 0)
            self.widths[key] = max(
                self.widths[key], max(len(row) for row in str(value).split("\n"))
            )
            height = max(height, str(value).count("\n"))
            row_format = self.row and self.format.text_wrap or self.format.bold
            self.worksheet.write(self.row, i, value, row_format)

        self.worksheet.set_row(self.row, 16 + height * 12)
        self.row += 1

    def get_header_values(self, fields, prefix="", key_prefix=""):
        if prefix:
            prefix += " - "
        if key_prefix:
            key_prefix += "_"
        for key, value in fields.items():
            if isinstance(value, ModelSerializer):
                yield from self.get_header_values(
                    value.get_fields(),
                    prefix + value.Meta.model._meta.verbose_name,
                    key_prefix + key,
                )
            else:
                self.header_keys.append(key_prefix + key)
                yield key_prefix + key, prefix + (getattr(value, "label", value) or key)

    def write_header(self, fields):
        self.write_values(list(self.get_header_values(fields)))

    def get_row_values(self, item, key_prefix=""):
        if key_prefix:
            key_prefix += "_"
        for key, value in item.items():
            if isinstance(value, OrderedDict):
                yield from self.get_row_values(value, key_prefix + key)
            else:
                yield key_prefix + key, value

    def write_row(self, item):
        self.write_values(self.get_row_values(item))

    def events_stats(self, queryset):
        self.add_worksheet("Účastníci a orgové akcí")
        participants = User.objects.filter(participated_in_events__event__in=queryset)
        organizers = User.objects.filter(events_where_was_organizer__in=queryset)
        main_organizers = User.objects.filter(
            events_where_was_as_main_organizer__in=queryset
        )

        self.write_header(
            dict(
                p="Účastníci",
                pe="Emaily",
                pc="Počet účastí",
                o="Orgové",
                oe="Emaily orgů",
                oc="Počet zorganizovaných akcí",
                m="Hlavní orgové",
                me="Emaily hlavních orgů",
                mc="Počet odvedených akcí",
            )
        )

        for line in zip_longest(
            *zip(*Counter(participants).most_common()),
            *zip(*Counter(organizers).most_common()),
            *zip(*Counter(main_organizers).most_common()),
            fillvalue="",
        ):
            row = []
            for item in line:
                if isinstance(item, User):
                    row += [item.get_name(), item.email or ""]
                else:
                    row += [item]

            row = {a: b for a, b in zip(self.header_keys, row)}
            self.write_row(row)

        for i in [0, 1, 3, 4, 6, 7]:
            self.worksheet.set_column(i, i, width=30)

        participants = (
            participants.annotate(count=Count("id")).distinct().order_by("-count")
        )
        organizers = (
            organizers.annotate(count=Count("id")).distinct().order_by("-count")
        )
        main_organizers = (
            main_organizers.annotate(count=Count("id")).distinct().order_by("-count")
        )
        self.from_queryset(participants, UserExportSerializer, "Účastníci")
        self.from_queryset(organizers, UserExportSerializer, "Organizátoři")
        self.from_queryset(main_organizers, UserExportSerializer, "Hlavní organizátoři")


executor = ThreadPoolExecutor(max_workers=1)


def send_later(request, result):
    try:
        file = result.result()
        emails.text(
            request.user.email,
            "Vygenerovaný export",
            "tumáš!",
            attachments=[
                {
                    "content_type": "xlsx",
                    "name": Path(file.name).name,
                    "data": b64encode(open(file.name, "rb").read()).decode(),
                }
            ],
        )
    except Exception as e:
        logging.exception(f"Error sending xlsx export to email: {e}")


@admin.action(description="Exportuj data")
def export_to_xlsx(model_admin, request, queryset):
    result = executor.submit(do_export_to_xlsx, queryset)
    for _ in range(40):
        sleep(1)
        if result.done():
            file = result.result()
            return FileResponse(open(file.name, "rb"))

    executor.submit(send_later, request, result)
    messages.warning(
        request,
        f"Buď tvůj export trvá dlouho, nebo se čeká na dokončení exportů ostatních, až bude hotov, pošlu ti ho na e-mail {request.user.email}",
    )


def export_to_xlsx_response(queryset):
    file = do_export_to_xlsx(queryset)
    return FileResponse(open(file.name, "rb"))


def do_export_to_xlsx(queryset):
    serializer_class = [
        s
        for s in [
            UserExportSerializer,
            EventExportSerializer,
            DonorExportSerializer,
            DonationExportSerializer,
            AdministrationUnitExportSerializer,
            EventApplicationExportSerializer,
            MembershipExportSerializer,
            EventFeedbackExportSerializer,
        ]
        if s.Meta.model is queryset.model
    ][0]
    writer = XLSXWriter(queryset.model._meta.verbose_name_plural)
    writer.from_queryset(queryset, serializer_class)
    if queryset.model is Event:
        writer.events_stats(queryset)

    file = writer.get_file()
    return file


def get_attendance_list_data(event, use_participants):
    organizers = list(event.other_organizers.all())
    if use_participants:
        rows = list(
            event.record.get_all_participants()
            if hasattr(event, "record")
            else event.other_organizers.all()
        )
    else:
        applications = list(
            EventApplication.objects.filter(
                state__in=["pending", "approved"], event_registration__event=event
            ).order_by("state", "created_at")
        )
        rows = organizers + applications

    for item in rows:
        address = getattr(item, "address", "")
        if not address and (applications_user := getattr(item, "user", None)):
            address = getattr(applications_user, "address", "")
        yield (
            item.first_name + " " + item.last_name,
            item.birthday and item.birthday.strftime("%d. %m. %Y"),
            address and f"{address.street}, {address.city}",
            address and address.zip_code,
            item.email,
            str(item.phone),
            item in organizers,
        )

    for i in range(max(10, len(rows) // 10)):
        yield 7 * ("",)


def get_attendance_list_rows(ws):
    for i in range(13):
        yield i + 17

    for i in range(25):
        yield i + 31

    start_i = 39
    start = 30
    rows = 26
    while True:
        for i in range(rows):
            source_row = start + i
            row = start + rows + i
            ws.row_dimensions[row] = copy(ws.row_dimensions[source_row])
            for col in "ABSCDEFGH":
                ws[f"{col}{row}"] = None
                ws[f"{col}{row}"]._style = copy(ws[f"{col}{source_row}"]._style)

            if i != 0:
                ws[f"A{row}"] = start_i
                start_i += 1

        start += rows

        for i in range(25):
            yield start + 1 + i


def get_attendance_list(event: Event, formatting, use_participants=False):
    if formatting == "xlsx":
        if use_participants:
            queryset = (
                hasattr(event, "record")
                and event.record.get_all_participants()
                or event.other_organizers.all()
            )
        else:
            queryset = EventApplication.objects.filter(
                state__in=["pending", "approved"], event_registration__event=event
            ).order_by("state", "created_at")

        return export_to_xlsx_response(queryset)

    assert formatting == "pdf"
    wb = openpyxl.load_workbook(
        join(BASE_DIR, "xlsx_export", "fixtures", "attendance_list_template.xlsx")
    )
    ws = wb.active

    ws["C2"] = event.name
    ws["C3"] = event.get_date()
    ws["C4"] = event.location.name
    ws["C5"] = ", ".join(au.abbreviation for au in event.administration_units.all())

    rows = get_attendance_list_rows(ws)
    data_rows = get_attendance_list_data(event, use_participants)
    for row, data in zip(rows, data_rows):
        for cell, value in zip("BCDEFG", data[:-1]):
            ws[f"{cell}{row}"] = value
        if data[-1]:
            for cell in "BCDEFGH":
                ws[f"{cell}{row}"].font = ws[f"{cell}{row}"].font.copy(bold=True)

    tmp_xlsx = NamedTemporaryFile(
        mode="w", suffix=".xlsx", newline="", encoding="utf8", prefix="attendance_list_"
    )
    tmp_html = NamedTemporaryFile(
        mode="w", suffix=".html", newline="", encoding="utf8", prefix="attendance_list_"
    )
    tmp_pdf = NamedTemporaryFile(
        mode="w", suffix=".pdf", newline="", encoding="utf8", prefix="attendance_list_"
    )

    wb.save(tmp_xlsx.name)

    xlsx2html(tmp_xlsx.name, tmp_html.name)

    pdfkit.from_file(
        tmp_html.name,
        tmp_pdf.name,
        {
            "page-size": "A4",
            "encoding": "UTF-8",
            "orientation": "Landscape",
            "title": "Landscape",
        },
    )

    return FileResponse(open(tmp_pdf.name, "rb"), as_attachment=True)


def export_files(event: Event):
    def mkdir(path):
        path.mkdir()
        return path

    file_name = f"Soubory {event.name}"
    with TemporaryDirectory() as tmp_dir:
        tmp_dir = Path(tmp_dir)
        if hasattr(event, "finance"):
            finance_path = mkdir(tmp_dir / _("models.EventFinance.name"))

            if event.finance.budget:
                copy2(event.finance.budget.path, finance_path)

            receipts_path = mkdir(
                finance_path / _("models.EventFinanceReceipt.name_plural")
            )
            for receipt in event.finance.receipts.all():
                copy2(receipt.receipt.path, receipts_path)

        if hasattr(event, "propagation"):
            propagation_path = mkdir(tmp_dir / _("models.EventPropagation.name"))

            images_path = mkdir(
                propagation_path / _("models.EventPropagationImage.name_plural")
            )
            for image in event.propagation.images.all():
                copy2(image.image.path, images_path)

        if hasattr(event, "record"):
            record_path = mkdir(tmp_dir / _("models.EventRecord.name"))

            attendance_list_pages_path = mkdir(
                record_path / _("models.EventAttendanceListPage.name_plural")
            )
            for attendance_list_page in event.record.attendance_list_pages.all():
                copy2(attendance_list_page.page.path, attendance_list_pages_path)

            photos_path = mkdir(record_path / _("models.EventPhoto.name_plural"))
            for photo in event.record.photos.all():
                copy2(photo.photo.path, photos_path)

        file = NamedTemporaryFile(
            mode="w", suffix=".zip", newline="", encoding="utf8", prefix=file_name + " "
        )

        with ZipFile(file.name, "w", ZIP_DEFLATED) as zip_file:
            for root, dirs, files in os.walk(tmp_dir):
                for _file in files:
                    zip_file.write(
                        os.path.join(root, _file),
                        os.path.relpath(os.path.join(root, _file), tmp_dir),
                    )

        os.chdir(tmp_dir)
        make_archive(file.name, "zip")

    return FileResponse(open(file.name, "rb"))


def get_donation_confirmation(donor):
    fixtures = Path(join(BASE_DIR, "xlsx_export", "fixtures"))

    # template is here https://docs.google.com/document/d/1mu5JAIUywyCz4LcbQ3BcQjHFAHbkhQ09/edit
    # and converted using https://pdf2png.com/
    background = Image.open(fixtures / "donation_confirmation_template.png")
    page = Image.new("RGB", background.size, (255, 255, 255))
    page.paste(background)
    page = page.resize((page.width * 3, page.height * 3))

    draw = ImageDraw.Draw(page)
    font = ImageFont.truetype(str(fixtures / "ARIAL.TTF"), size=3 * 34)

    today = date.today()
    year = today.year - 1 if today.month < 6 else today.year
    created_at = today
    if today.month < 6:
        created_at = min(today, date(today.year, 1, 31))

    text = date_format(created_at)
    text_params = dict(fill=(0, 0, 0), font=font, spacing=1.5 * 34)
    draw.text((1244 * 3, 379 * 3 - 1), text, **text_params)

    total = sum(
        donation.amount for donation in donor.donations.filter(donated_at__year=year)
    )
    assert total > 0, f"{donor} za poslední rok nic nedaroval"

    pronoun = donor.user.pronoun and donor.user.pronoun.slug
    pronoun_texts = ["pan/slečna/paní", "poskytl/a"]
    if pronoun == "man":
        pronoun_texts = ["pan", "poskytl"]
    if pronoun == "woman":
        pronoun_texts = ["paní", "poskytla"]

    text = (
        f"Potvrzujeme, že {pronoun_texts[0]} {donor.user.first_name} {donor.user.last_name}, "
        f"trvale bytem {donor.user.address}, "
        f"v roce {year} {pronoun_texts[1]} dar Hnutí Brontosaurus ve výši {total} Kč. "
    )

    text += "Tento dar byl v souladu se Zákonem č. 586/1992 Sb., o daních z příjmů, ve znění pozdějších předpisů dle § 15 odst. 1, poskytnut na podporu mládeže a ekologické účely."
    text = text_into_lines(draw, font, text, 4000)
    draw.text((176 * 3, 600 * 3 - 1), text, **text_params)

    page = page.resize((page.width // 3, page.height // 3), Image.ANTIALIAS)
    tmp_pdf = NamedTemporaryFile(
        mode="w", suffix=".pdf", newline="", encoding="utf8", prefix="attendance_list_"
    )
    page.save(tmp_pdf.name)
    return open(tmp_pdf.name, "rb"), year


def get_donation_points(donation_points):
    wb = openpyxl.Workbook()
    ws = wb.active

    ws.cell(1, 1).value = "Body dotací"
    ws.cell(2, 1).value = "od"
    ws.cell(2, 2).value = str(donation_points.since)
    ws.cell(2, 3).value = "do"
    ws.cell(2, 4).value = str(donation_points.till)

    for r, row in enumerate(donation_points.get_rows()):
        for c, cell in enumerate(row):
            ws.cell(4 + r, 1 + c).value = cell

    tmp_file = NamedTemporaryFile(
        mode="w",
        suffix=".xlsx",
        newline="",
        encoding="utf8",
        prefix="donation_points_",
    )
    wb.save(tmp_file.name)
    return open(tmp_file.name, "rb")
